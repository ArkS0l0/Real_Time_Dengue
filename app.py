import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import requests
import os
import folium
from folium.plugins import HeatMap, Fullscreen, MiniMap, MeasureControl, MousePosition
from streamlit_folium import st_folium
import io
from datetime import date
from openpyxl import load_workbook
import uuid
from branca.element import MacroElement
from jinja2 import Template

# Query-param compatibility helpers
import urllib.parse

def get_query_params_safe():
    """Return query params using the stable API when available, with fallbacks.

    Returns a plain dict (keys -> list-of-values) for compatibility.
    """
    # Prefer the stable property first
    try:
        return dict(st.query_params)
    except Exception:
        pass
    # Fallback to older experimental API if present
    try:
        return st.experimental_get_query_params()
    except Exception:
        return {}

def set_query_params_safe(params: dict):
    try:
        # Prefer the stable API if available
        try:
            st.set_query_params(**params)
            return
        except Exception:
            pass
        # Try the older experimental API
        try:
            st.experimental_set_query_params(**params)
            return
        except Exception:
            pass
        # Fallback: client-side set via window.location.search
        qs = '&'.join(f"{urllib.parse.quote_plus(str(k))}={urllib.parse.quote_plus(str(v))}" for k, v in params.items())
        st.markdown(f"<script>window.location.search = '?{qs}';</script>", unsafe_allow_html=True)
    except Exception:
        # Last resort: do nothing
        return

def clear_query_params_safe():
    try:
        # Try stable API
        try:
            st.set_query_params()
            return
        except Exception:
            pass
        # Fallback to experimental API
        try:
            st.experimental_set_query_params()
            return
        except Exception:
            pass
        # Final fallback: replace URL client-side and reload
        st.markdown("""<script>window.history.replaceState(null,'', window.location.pathname); window.location.reload();</script>""", unsafe_allow_html=True)
    except Exception:
        return

def reload_safe():
    try:
        st.experimental_rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            clear_query_params_safe()

    
# Title
st.title("🦟 Dengue Risk Prediction System")
st.markdown("**Real-time dengue danger assessment with AI-powered mitigation recommendations**")

# Sidebar inputs
st.sidebar.header("📝 Input Parameters")
area = st.sidebar.text_input("Area name (e.g., Woodlands, Ang Mo Kio)", "Woodlands", key='area_input')
construction = st.sidebar.checkbox("Construction sites present?", True, key='construction_chk')

# Navigation will be added after inputs and scoring so map can reflect computed risk

# Weather section
st.sidebar.header("🌤️ Weather Data")
weather_source = st.sidebar.radio("Weather source:", ["Manual Entry", "OpenWeatherMap API"], key='weather_source')

humidity = 70
temperature = 30
rainfall = 0.0

if weather_source == "OpenWeatherMap API":
    api_key = st.sidebar.text_input("OpenWeatherMap API Key", type="password", placeholder="Enter your API key", key='owm_api_key')
    if api_key and st.sidebar.button("Fetch Weather", key='fetch_weather'):
        try:
            # Geocoding
            geo_url = f'http://api.openweathermap.org/geo/1.0/direct?q={area}&limit=1&appid={api_key}'
            geo_response = requests.get(geo_url, timeout=10)
            geo = geo_response.json()
            if geo:
                lat, lon = geo[0]['lat'], geo[0]['lon']
                # Weather
                weather_url = f'https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&appid={api_key}'
                w_response = requests.get(weather_url, timeout=10)
                w = w_response.json()['main']
                st.sidebar.success("✓ Weather data fetched!")
                humidity = w['humidity']
                temperature = w['temp']
                rainfall = 0.0
        except Exception as e:
            st.sidebar.error(f"Error fetching weather: {e}")
else:
    humidity = st.sidebar.slider("Humidity (%)", 0, 100, 70, key='humidity')
    temperature = st.sidebar.slider("Temperature (°C)", 15, 45, 30, key='temperature')
    rainfall = st.sidebar.slider("Recent rainfall (mm)", 0.0, 100.0, 0.0, key='rainfall')

# Dengue cases section
st.sidebar.header("🔴 Dengue Data")
cases_source = st.sidebar.radio("Cases source:", ["Local CSV", "Manual Entry"], key='cases_source')

historical_cases = 39
active_clusters = 5

if cases_source == "Local CSV":
    if os.path.exists('dengue_cases.csv'):
        try:
            df = pd.read_csv('dengue_cases.csv')
            # Make matching robust: cast first column to str before lowercasing
            area_match = df[df.iloc[:, 0].astype(str).str.lower() == area.lower()]
            if not area_match.empty:
                historical_cases = int(area_match.iloc[0, 1])
                active_clusters = int(area_match.iloc[0, 2]) if len(area_match.columns) > 2 else 0
                st.sidebar.success(f"✓ Loaded {historical_cases} cases for {area}")
            else:
                historical_cases = st.sidebar.number_input("Dengue cases last week", 0, 1000, 39, key='cases_last_week')
                active_clusters = st.sidebar.number_input("Active clusters", 0, 50, 5, key='active_clusters')
        except:
            historical_cases = st.sidebar.number_input("Dengue cases last week", 0, 1000, 39, key='cases_last_week')
            active_clusters = st.sidebar.number_input("Active clusters", 0, 50, 5, key='active_clusters')
    else:
        st.sidebar.info("dengue_cases.csv not found. Using manual entry.")
        historical_cases = st.sidebar.number_input("Dengue cases last week", 0, 1000, 39, key='cases_last_week')
        active_clusters = st.sidebar.number_input("Active clusters", 0, 50, 5, key='active_clusters')
else:
    historical_cases = st.sidebar.number_input("Dengue cases last week", 0, 1000, 39)
    active_clusters = st.sidebar.number_input("Active clusters", 0, 50, 5)

# Sidebar heatmap uploader (stores uploaded CSV path in session state)
st.sidebar.header("🗺️ Heatmap Upload")
uploaded_heatmap = st.sidebar.file_uploader("Upload CSV with lat/lon for heatmap", type=['csv'], key='sidebar_heatmap')
if uploaded_heatmap is not None:
    try:
        df_h = pd.read_csv(uploaded_heatmap)
        upload_path = 'uploaded_heatmap.csv'
        df_h.to_csv(upload_path, index=False)
        st.sidebar.success('Heatmap CSV uploaded')
        st.session_state['heatmap_path'] = upload_path
    except Exception as e:
        st.sidebar.error(f'Upload error: {e}')
else:
    # Prefer previously uploaded file or sample if present
    if 'heatmap_path' not in st.session_state and os.path.exists('uploaded_heatmap.csv'):
        st.session_state['heatmap_path'] = 'uploaded_heatmap.csv'
    elif 'heatmap_path' not in st.session_state and os.path.exists('sample_heatmap.csv'):
        st.session_state['heatmap_path'] = 'sample_heatmap.csv'

if st.sidebar.button('Clear heatmap upload', key='clear_heatmap'):
    if 'heatmap_path' in st.session_state:
        st.session_state.pop('heatmap_path', None)
    if os.path.exists('uploaded_heatmap.csv'):
        try:
            os.remove('uploaded_heatmap.csv')
        except Exception:
            pass
    st.sidebar.success('Cleared heatmap upload')

# SCORING ALGORITHM
score = 0.0
score += 2 if construction else 0
score += max(0, min(2, (float(humidity) - 60) / 7.5)) if humidity > 60 else 0
if temperature >= 30:
    score += 1.5
elif temperature >= 26:
    score += 0.8
if rainfall >= 50:
    score += 2
elif rainfall >= 20:
    score += 1
if historical_cases >= 20:
    score += 3
elif historical_cases >= 5:
    score += 1
if active_clusters >= 3:
    score += 1

score = int(min(10, round(score)))

# Danger level
if score >= 6:
    danger = "High"
    color = "🔴"
elif score >= 3:
    danger = "Medium"
    color = "🟡"
else:
    danger = "Low"
    color = "🟢"

# Provide a full-page map route via query params (page=map). If present, render only the large map.
params = get_query_params_safe()
if params.get('page', [''])[0] == 'map':
    map_area = params.get('area', [area])[0]
    try:
        geo_resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": map_area, "format": "json", "limit": 1},
            headers={"User-Agent": "dengue-app/1.0"},
            timeout=10,
        )
        geo = geo_resp.json()
        if geo:
            lat = float(geo[0]["lat"])
            lon = float(geo[0]["lon"])
            # Large, scalable map view with enhanced controls
            m_full = folium.Map(location=[lat, lon], zoom_start=13, control_scale=True, prefer_canvas=True)
            # Add useful plugins
            Fullscreen(position='topright').add_to(m_full)
            MiniMap(toggle_display=True, position='bottomright').add_to(m_full)
            MeasureControl(position='topleft', primary_length_unit='meters').add_to(m_full)
            MousePosition().add_to(m_full)
            marker_color = {'High': 'red', 'Medium': 'orange', 'Low': 'green'}.get(danger, 'blue')
            folium.CircleMarker(
                location=[lat, lon],
                radius=max(8, int(score * 3)),
                color=marker_color,
                fill=True,
                fill_color=marker_color,
                fill_opacity=0.6,
                popup=f"{map_area}: {danger} (Score {score})",
            ).add_to(m_full)
            folium.LayerControl().add_to(m_full)
            st.markdown(f"### 🗺️ Full Map — {map_area}")
            if st.button("Back to dashboard", key='back_dashboard'):
                # Clear query params and reload using compatibility helpers
                clear_query_params_safe()
                reload_safe()
            st_folium(m_full, width=1200, height=900)
        else:
            st.error(f"Could not geocode area: {map_area}")
    except Exception as e:
        st.error(f"Map page error: {e}")
    st.stop()

# Display results inside tabs
tabs = st.tabs(["Dashboard", "Embedded Map", "Heatmap", "Newsletter"])

with tabs[0]:
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Risk Score", f"{score}/10", delta=None)
        
    with col2:
        st.metric("Danger Level", f"{color} {danger}")
        
    with col3:
        st.metric("Active Clusters", active_clusters)

    # Action recommendations
    action_map = {
        'High': 'Immediate — Enhanced surveillance, fogging where appropriate, community alerts, and intensive source reduction',
        'Medium': 'Targeted inspections, community outreach, larvicide in hotspots, weekly monitoring',
        'Low': 'Routine monitoring and public education; update weekly'
    }

    st.subheader("📋 Recommended Action Level")
    st.info(action_map[danger])

    # Mitigation suggestions
    st.subheader("💡 Tailored Mitigation Suggestions")
    suggestions = []
    if construction:
        suggestions.append("Inspect construction sites and drains; enforce water management")
    if humidity >= 75:
        suggestions.append("Increase stagnant water checks (humid conditions favor mosquitoes)")
    if temperature >= 30:
        suggestions.append("Focus on outdoor breeding spots; adult mosquitoes more active")
    if rainfall >= 50:
        suggestions.append("Clear rainwater from containers and debris")
    if historical_cases >= 5:
        suggestions.append("Conduct targeted house-to-house inspections in clusters")
    if active_clusters >= 3:
        suggestions.append("Set up temporary cluster response teams")

    if suggestions:
        for i, s in enumerate(suggestions, 1):
            st.success(f"{i}. {s}")
    else:
        st.success("Maintain routine monitoring and public education")

    # Visualization: Risk score trend
    st.subheader("📈 4-Week Risk Trend")
    weeks = [datetime.today() - timedelta(weeks=i) for i in range(3, -1, -1)]
    np.random.seed(42)
    past_scores = np.clip(np.random.normal(score, 1, 4).astype(int), 0, 10)

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(weeks, past_scores, marker='o', linewidth=2, markersize=8, color='#FF6B6B')
    ax.fill_between(weeks, past_scores, alpha=0.3, color='#FF6B6B')
    ax.set_ylabel('Risk Score', fontsize=12)
    ax.set_title(f'Dengue Risk Trend - {area}', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    st.pyplot(fig, width='stretch')

    # Danger level circle
    st.subheader("🎯 Danger Indicator")
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        fig, ax = plt.subplots(figsize=(6, 6))
        color_map = {'High': '#FF6B6B', 'Medium': '#FFD93D', 'Low': '#6BCB77'}
        circle = plt.Circle((0.5, 0.5), 0.4, color=color_map[danger], alpha=0.8)
        ax.add_patch(circle)
        ax.text(0.5, 0.5, f'{danger}\n(Score: {score})', ha='center', va='center', 
                fontsize=24, fontweight='bold', color='white')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis('off')
        st.pyplot(fig, width='stretch')

    # Input summary
    st.subheader("📊 Input Summary")
    input_df = pd.DataFrame({
        'Parameter': ['Area', 'Construction Sites', 'Humidity (%)', 'Temperature (C)', 'Recent Rainfall (mm)', 'Cases (Last Week)', 'Active Clusters'],
        'Value': [area, 'Yes' if construction else 'No', humidity, temperature, rainfall, historical_cases, active_clusters]
    })
    # Ensure consistent types for Streamlit/Arrow serialization
    input_df['Value'] = input_df['Value'].astype(str)
    st.table(input_df)

with tabs[1]:
    # When this tab is selected, redirect to the full-page map route via query params
    params = get_query_params_safe()
    if params.get('page', [''])[0] != 'map':
        # Set query params via compatibility wrapper (will use client-side JS fallback if needed)
        set_query_params_safe({'page': 'map', 'area': area})
        st.stop()

    # Map view: Geocode area with Nominatim (OpenStreetMap) and show hotspot marker
    try:
        geo_resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": area, "format": "json", "limit": 1},
            headers={"User-Agent": "dengue-app/1.0"},
            timeout=10,
        )
        geo = geo_resp.json()
        if geo:
            lat = float(geo[0]["lat"]) 
            lon = float(geo[0]["lon"]) 
            m = folium.Map(location=[lat, lon], zoom_start=13, control_scale=True, prefer_canvas=True)
            Fullscreen(position='topright').add_to(m)
            MiniMap(toggle_display=True, position='bottomright').add_to(m)
            MeasureControl(position='topleft', primary_length_unit='meters').add_to(m)
            MousePosition().add_to(m)
            # Color by danger level
            marker_color = {'High': 'red', 'Medium': 'orange', 'Low': 'green'}.get(danger, 'blue')
            folium.CircleMarker(
                location=[lat, lon],
                radius=max(8, int(score * 3)),
                color=marker_color,
                fill=True,
                fill_color=marker_color,
                fill_opacity=0.6,
                popup=f"{area}: {danger} (Score {score})",
            ).add_to(m)
            st.subheader(f"🗺️ Dengue hotspots — {area}")
            st_folium(m, width=700, height=500)
        else:
            st.error(f"Could not geocode area: {area}")
    except Exception as e:
        st.error(f"Map error: {e}")

with tabs[2]:
    st.subheader("🌡️ Heatmap of Dengue Cases")
    # Try to load lat/lon data from dengue_cases.csv
    # Prefer uploaded heatmap CSV if present in session state
    heatmap_file = st.session_state.get('heatmap_path') if 'heatmap_path' in st.session_state else None
    if heatmap_file and os.path.exists(heatmap_file):
        source_path = heatmap_file
    elif os.path.exists('dengue_cases.csv'):
        source_path = 'dengue_cases.csv'
    elif os.path.exists('sample_heatmap.csv'):
        source_path = 'sample_heatmap.csv'
    else:
        source_path = None

    if source_path:
        try:
            df_cases = pd.read_csv(source_path)
            # look for lat/lon columns
            lat_cols = [c for c in df_cases.columns if c.lower() in ('lat','latitude','y','latitud')]
            lon_cols = [c for c in df_cases.columns if c.lower() in ('lon','lng','longitude','x','long')]
            if lat_cols and lon_cols:
                lat_col = lat_cols[0]
                lon_col = lon_cols[0]
                heat_data = df_cases[[lat_col, lon_col]].dropna().values.tolist()
                if heat_data:
                    center = [float(heat_data[0][0]), float(heat_data[0][1])]
                    m2 = folium.Map(location=center, zoom_start=12, control_scale=True, prefer_canvas=True)
                    Fullscreen(position='topright').add_to(m2)
                    MiniMap(toggle_display=True, position='bottomright').add_to(m2)
                    MeasureControl(position='topleft', primary_length_unit='meters').add_to(m2)
                    MousePosition().add_to(m2)
                    HeatMap(heat_data, radius=15, blur=10, max_zoom=13).add_to(m2)
                    # Add simple legend
                    legend_html = '''
                        <div style="position: fixed; 
                                    bottom: 50px; left: 50px; width: 150px; height: 120px; 
                                    background-color: white; border:2px solid grey; z-index:9999; font-size:14px; padding:8px; border-radius:8px;">
                        <b>Heatmap legend</b><br>
                        <i style="background:rgba(255,0,0,0.6);width:12px;height:12px;display:inline-block;margin-right:6px;"></i> High<br>
                        <i style="background:rgba(255,165,0,0.6);width:12px;height:12px;display:inline-block;margin-right:6px;"></i> Medium<br>
                        <i style="background:rgba(0,128,0,0.6);width:12px;height:12px;display:inline-block;margin-right:6px;"></i> Low
                        </div>'''
                    m2.get_root().html.add_child(folium.Element(legend_html))
                    st_folium(m2, width=800, height=600)
                else:
                    st.info(f"No lat/lon rows found in {source_path}")
            else:
                st.info(f"{source_path} found but no lat/lon columns detected. Please upload a CSV with latitude/longitude columns named 'lat'/'lon' or similar.")
        except Exception as e:
            st.error(f"Error loading {source_path}: {e}")
    else:
        st.info("dengue_cases.csv not found. Upload a CSV with lat/lon to enable heatmap.")

    # Allow user to upload a CSV for heatmap
    uploaded = st.file_uploader("Upload CSV with lat/lon for heatmap", type=['csv'], key='tab_heatmap')
    if uploaded is not None:
        try:
            df_up = pd.read_csv(uploaded)
            lat_cols = [c for c in df_up.columns if c.lower() in ('lat','latitude','y','latitud')]
            lon_cols = [c for c in df_up.columns if c.lower() in ('lon','lng','longitude','x','long')]
            if lat_cols and lon_cols:
                lat_col = lat_cols[0]
                lon_col = lon_cols[0]
                heat_data = df_up[[lat_col, lon_col]].dropna().values.tolist()
                center = [float(heat_data[0][0]), float(heat_data[0][1])]
                m3 = folium.Map(location=center, zoom_start=12, control_scale=True, prefer_canvas=True)
                Fullscreen(position='topright').add_to(m3)
                MiniMap(toggle_display=True, position='bottomright').add_to(m3)
                MeasureControl(position='topleft', primary_length_unit='meters').add_to(m3)
                MousePosition().add_to(m3)
                HeatMap(heat_data, radius=15, blur=10, max_zoom=13).add_to(m3)
                legend_html = '''
                    <div style="position: fixed; 
                                bottom: 50px; left: 50px; width: 150px; height: 120px; 
                                background-color: white; border:2px solid grey; z-index:9999; font-size:14px; padding:8px; border-radius:8px;">
                    <b>Heatmap legend</b><br>
                    <i style="background:rgba(255,0,0,0.6);width:12px;height:12px;display:inline-block;margin-right:6px;"></i> High<br>
                    <i style="background:rgba(255,165,0,0.6);width:12px;height:12px;display:inline-block;margin-right:6px;"></i> Medium<br>
                    <i style="background:rgba(0,128,0,0.6);width:12px;height:12px;display:inline-block;margin-right:6px;"></i> Low
                    </div>'''
                m3.get_root().html.add_child(folium.Element(legend_html))
                st_folium(m3, width=800, height=600)
            else:
                st.error("Uploaded CSV does not contain recognizable lat/lon columns.")
        except Exception as e:
            st.error(f"Error processing uploaded CSV: {e}")

with tabs[3]:
    st.subheader("📰 Newsletter & Alerts")
    newsletter_path = 'newsletter.csv'
    if os.path.exists(newsletter_path):
        try:
            news_df = pd.read_csv(newsletter_path)
            st.table(news_df.sort_values('timestamp', ascending=False).head(10))
        except Exception:
            st.info('No newsletter entries yet')
    else:
        st.info('No newsletter entries yet')

    st.markdown('Add an announcement (these are saved to newsletter.csv)')
    ann = st.text_area('Announcement', '', height=120, key='announcement')
    ann_level = st.selectbox('Severity', ['Info','Warning','Alert'], key='announcement_level')
    if st.button('Publish Announcement', key='publish_announcement') and ann.strip():
        new_row = {'id': str(uuid.uuid4()), 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'level': ann_level, 'message': ann}
        try:
            if os.path.exists(newsletter_path):
                dfn = pd.read_csv(newsletter_path)
                dfn = pd.concat([pd.DataFrame([new_row]), dfn], ignore_index=True)
            else:
                dfn = pd.DataFrame([new_row])
            dfn.to_csv(newsletter_path, index=False)
            st.success('Announcement published')
        except Exception as e:
            st.error(f'Error saving announcement: {e}')

    # Export weekly cases to Excel workbook (3D sheets: each sheet = iso-week)
    st.markdown("**Export weekly cases**")
    week_label = date.today().isocalendar()
    sheet_name = f"{week_label[0]}-W{week_label[1]:02d}"
    if st.button("Save current area to weekly Excel", key='save_weekly'):
        row = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'area': area,
            'construction': 'Yes' if construction else 'No',
            'humidity_pct': humidity,
            'temperature_c': temperature,
            'rainfall_mm': rainfall,
            'historical_cases': historical_cases,
            'active_clusters': active_clusters,
            'danger_score': score,
            'danger_level': danger,
        }
        excel_path = 'weekly_cases.xlsx'
        df_row = pd.DataFrame([row])
        try:
            if os.path.exists(excel_path):
                book = load_workbook(excel_path)
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # If sheet exists, append; otherwise create
                    if sheet_name in book.sheetnames:
                        existing = pd.read_excel(excel_path, sheet_name=sheet_name)
                        combined = pd.concat([existing, df_row], ignore_index=True)
                        combined.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        df_row.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df_row.to_excel(writer, sheet_name=sheet_name, index=False)
            st.success(f"Saved to {excel_path} sheet {sheet_name}")
        except Exception as e:
            st.error(f"Error saving Excel: {e}")

# Input summary
st.subheader("📊 Input Summary")
input_df = pd.DataFrame({
    'Parameter': ['Area', 'Construction Sites', 'Humidity (%)', 'Temperature (C)', 'Recent Rainfall (mm)', 'Cases (Last Week)', 'Active Clusters'],
    'Value': [area, 'Yes' if construction else 'No', humidity, temperature, rainfall, historical_cases, active_clusters]
})
# Ensure consistent types for Streamlit/Arrow serialization
input_df['Value'] = input_df['Value'].astype(str)
st.table(input_df)

# Save to CSV
st.sidebar.header("💾 Export")
if st.sidebar.button("Save Report", key='save_report'):
    report_row = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'area': area,
        'construction': 'Yes' if construction else 'No',
        'humidity_pct': humidity,
        'temperature_c': temperature,
        'rainfall_mm': rainfall,
        'historical_cases': historical_cases,
        'active_clusters': active_clusters,
        'danger_score': score,
        'danger_level': danger,
        'recommended_action': action_map[danger],
    }
    
    try:
        if os.path.exists('dengue_report.csv'):
            df_existing = pd.read_csv('dengue_report.csv')
            df_new = pd.DataFrame([report_row])
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            df_combined.to_csv('dengue_report.csv', index=False)
        else:
            pd.DataFrame([report_row]).to_csv('dengue_report.csv', index=False)
        
        st.sidebar.success("Report saved to dengue_report.csv")
    except Exception as e:
        st.sidebar.error(f"Error saving report: {e}")

# Footer
st.markdown("---")
footer_text = (
    "**Dengue Risk Prediction System**\n\n"
    "This tool is for public health planning purposes. Always consult official health authorities for guidance.\n\n"
    "**Data Sources:**\n"
    "- Weather: OpenWeatherMap API\n"
    "- Dengue cases: Local CSV or manual entry\n"
    "- Methodology: WHO & CDC guidelines\n\n"
    "**License:** MIT (Open Source) | **Repository:** [GitHub](https://github.com)\n"
)
st.markdown(footer_text)
